import io
import os
import re
import glob
import requests
import pandas as pd
from datetime import datetime
import openpyxl

# =======================================================
FOLDER_DIR = "."
# =======================================================

def get_web_data_and_date(url):
    """提取网页数据，统一返回 YYYY-MM-DD"""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8"
    }
    session = requests.Session()
    response = session.get(url, headers=headers, timeout=20)
    response.raise_for_status()
    response.encoding = 'utf-8'
    html_text = response.text

    date_match = re.search(r"更新[^\d]{0,20}(\d{4}[-/]\d{1,2}[-/]\d{1,2})", html_text)
    if date_match:
        real_date = date_match.group(1).replace("/", "-")
    else:
        real_date = datetime.now().strftime("%Y-%m-%d")

    html_data = io.StringIO(html_text)
    tables = pd.read_html(html_data)
    if not tables:
        return [], real_date

    valid_tables = [t for t in tables if len(t.columns) >= 6]
    return valid_tables, real_date


def clean_string(s):
    s = str(s).upper().replace(" ", "").replace("*", "X").replace("（", "(").replace("）", ")").strip()
    s = s.replace("($USD)", "").replace("(USD)", "")
    return s


def get_row_data_dict(tables_list, df_history):
    row_data = {}
    for df_today in tables_list:
        for _, row in df_today.iterrows():
            item_name = str(row.iloc[0]).strip()
            if not item_name or item_name == 'nan' or '日期' in item_name:
                continue

            clean_item = clean_string(item_name)

            for i, col in enumerate(df_history.columns):
                col_group = col[0]
                if 'Unnamed' in str(col_group):
                    continue
                clean_group = clean_string(col_group)

                if clean_item == clean_group or clean_item in clean_group or clean_group in clean_item:
                    indicator = col[1]
                    val = None
                    try:
                        if indicator == '日高点':
                            val = str(row.iloc[1]).strip()
                        elif indicator == '日低点':
                            val = str(row.iloc[2]).strip()
                        elif indicator == '盘高点':
                            val = str(row.iloc[3]).strip()
                        elif indicator == '盘低点':
                            val = str(row.iloc[4]).strip()
                        elif indicator == '盘平均':
                            val = str(row.iloc[5]).strip()
                    except:
                        continue

                    if val and val != 'nan' and val != 'None' and val != '-':
                        row_data[i] = val
    return row_data


def get_real_max_row(ws):
    """
    【终极防线】：绝对物理扫描。无视任何 Excel 缓存、幽灵边框、被破坏的维度标签。
    只认实实在在的文字！从根本上粉碎“代码只能用一次”的覆盖 Bug！
    """
    max_r = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "" and str(cell.value).strip().lower() != 'nan':
                if cell.row > max_r:
                    max_r = cell.row
    return max_r


def write_to_excel_cell(ws, df_history, row_data_dict, real_date_str):
    standard_web_date = datetime.strptime(real_date_str, "%Y-%m-%d").strftime("%Y-%m-%d")
    
    # 1. 🛡️ 启动全盘物理扫描，获取绝对真实的最后一行
    real_max = get_real_max_row(ws)
    
    # 2. 🔍 提取最后一行的真实日期
    last_date_val = ws.cell(row=real_max, column=1).value
    last_date_str = ""
    if last_date_val:
        if isinstance(last_date_val, datetime):
            last_date_str = last_date_val.strftime("%Y-%m-%d")
        else:
            try:
                clean_str = str(last_date_val).strip().split()[0].replace("/", "-")
                last_date_str = datetime.strptime(clean_str, "%Y-%m-%d").strftime("%Y-%m-%d")
            except:
                last_date_str = str(last_date_val)

    # 3. 🔀 刚性判定写入行
    if last_date_str == standard_web_date:
        print(f"\n 📅 工作表【{ws.title}】：检测到网页日期 ({standard_web_date}) 与最后一行相同 -> 执行【原位热更新】，覆盖第 {real_max} 行。")
        target_row = real_max
    else:
        target_row = real_max + 1
        print(f"\n 📅 工作表【{ws.title}】：检测到全新开盘日期 ({standard_web_date}) -> 全视之眼锁定安全空位，追加至第 {target_row} 行。")

    # 4. ✍️ 横向填满所有属于“日期”的单元格
    dt_obj = datetime.strptime(standard_web_date, "%Y-%m-%d")
    formatted_date = f"{dt_obj.year}/{dt_obj.month}/{dt_obj.day}"

    for c_idx, col in enumerate(df_history.columns):
        if col[1] == '日期':
            ws.cell(row=target_row, column=c_idx + 1).value = formatted_date

    # 5. 💾 写入各维度的具体价格
    for col_idx, value in row_data_dict.items():
        if value is not None and str(value) != 'nan':
            try:
                ws.cell(row=target_row, column=col_idx + 1).value = float(value)
            except ValueError:
                ws.cell(row=target_row, column=col_idx + 1).value = value

    return True


def main():
    print("\n==================================================")

    # ---------------- 自动化模块：寻找文件 ---------------- #
    print("🔍 [0/4] 自动雷达：正在定位本地最新的 Excel 追踪表...")
    search_pattern = os.path.join(FOLDER_DIR, "内存价格每日追踪_*.xlsx")
    matched_files = glob.glob(search_pattern)

    if not matched_files:
        print(f"❌ 错误：在 {FOLDER_DIR} 下找不到名字包含【内存价格每日追踪_】的表格！")
        return

    current_excel = max(matched_files)
    print(f"📂 成功锁定唯一历史火种文件：【{os.path.basename(current_excel)}】\n")
    # -------------------------------------------------- #

    print("🚀 [1/4] 开始同步 TrendForce 官方数据...")
    try:
        tables_dram, dram_date = get_web_data_and_date("https://www.trendforce.cn/price")
        tables_flash, flash_date = get_web_data_and_date("https://www.trendforce.cn/price/flash")

        print("\n📦 [2/4] 正在进行多级表头数据结构对齐...")
        df_dram_hist = pd.read_excel(current_excel, sheet_name='DRAM Spot Price', header=[0, 1], engine='openpyxl')
        df_flash_hist = pd.read_excel(current_excel, sheet_name='NAND Flash', header=[0, 1], engine='openpyxl')

        dram_row_dict = get_row_data_dict(tables_dram, df_dram_hist)
        flash_row_dict = get_row_data_dict(tables_flash, df_flash_hist)

        print("💾 [3/4] 正在安全注入 Excel 底层单元格...")
        wb = openpyxl.load_workbook(current_excel)
        write_to_excel_cell(wb['DRAM Spot Price'], df_dram_hist, dram_row_dict, dram_date)
        write_to_excel_cell(wb['NAND Flash'], df_flash_hist, flash_row_dict, flash_date)
        wb.save(current_excel)
        wb.close()

        # ---------------- 自动化模块：修改文件名 ---------------- #
        print("🏷️ [4/4] 正在执行智能文件升级与迭代...")
        today_compact = datetime.now().strftime("%Y%m%d")  
        new_filename = f"内存价格每日追踪_{today_compact}.xlsx"
        new_filepath = os.path.join(FOLDER_DIR, new_filename)

        if os.path.abspath(current_excel) != os.path.abspath(new_filepath):
            if os.path.exists(new_filepath):
                os.replace(current_excel, new_filepath)
            else:
                os.rename(current_excel, new_filepath)
            print(f"✅ 文件已全自动进化更名为 ──> 【{new_filename}】")
        else:
            print(f"✅ 数据已在今日最新表格内完成原位热更新 ──> 【{new_filename}】")
        # -------------------------------------------------- #

        print("\n✨✨✨ 微信/网盘全自动托管模式运行成功！历史链路无缝闭环！ ✨✨✨")
        print("==========================================================================\n")

    except PermissionError:
        print(f"\n❌ 写入失败！系统报错：Permission denied")
    except Exception as e:
        print(f"❌ 运行出错了: {e}")


if __name__ == "__main__":
    main()
